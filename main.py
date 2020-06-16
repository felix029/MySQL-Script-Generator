import sys
import getopt
from datetime import datetime

from openpyxl import load_workbook

def main():
    table       = ''
    path        = ''
    filename    = ''
    script      = ''


    options, remainder = getopt.getopt(sys.argv[1:], '', ['table=', 'path='])

    for opt, arg in options:
        
        if opt == '--table':
            table   = arg

        if opt == '--path':
            path    = arg
        

    # Read given xlsx
    wb = load_workbook(filename = path)
    ws = wb.active

    row_count       = ws.max_row
    column_count    = ws.max_column

    #Create a filename for the script
    now = datetime.now()
    filename = table + "_" + now.strftime("%Y-%m-%d") + ".sql"

    #Create a new file
    f = open(filename, "w")

    script += "INSERT INTO " + table 

    #Iterate through the xlsx file and create the SQL file
    rows = ws.rows
    for i, row in enumerate(rows, 1):
        for j, cell in enumerate(row, 1):
            #First row of the xlsx, table columns names
            if i == 1:
                if j == 1:
                    script += "(`" + str(cell.value) + "`, "
                elif j == column_count:
                    script += "`" + str(cell.value) + "`) VALUES "
                else:
                    script += "`"+ str(cell.value) + "`, "

            #Last row of the xlsx, ; is needed at the end of the file
            elif i == row_count:
                if str(cell.value) == "NULL" or str(cell.value) == "null":
                    if j == 1:
                        script += "(" + str(cell.value) + ", " 
                    elif j == column_count:
                        script += str(cell.value) + ");"
                    else:
                        script += str(cell.value) + ", "

                else:
                    if j == 1:
                        script += "('" + str(cell.value) + "', " 
                    elif j == column_count:
                        script += "'" + str(cell.value) + "');"
                    else:
                        script += "'" + str(cell.value) + "', "

            #Normal rows
            else:
                if str(cell.value) == "NULL" or str(cell.value) == "null":
                    if j == 1:
                        script += "(" + str(cell.value) + ", "
                    elif j == column_count:
                        script += str(cell.value) + "), "
                    else:
                        script += str(cell.value) + ", "

                else:
                    if j == 1:
                        script += "('" + str(cell.value) + "', "
                    elif j == column_count:
                        script += "'" + str(cell.value) + "'), "
                    else:
                        script += "'" + str(cell.value) + "', "
    
    f.write(script)
    f.close()

    return 0

if __name__ == '__main__':
    sys.exit(main())
